"""测试 common.storage 深模块及各子项目薄 adapter。"""

import csv
from pathlib import Path

from common.storage import append_csv, merge_csv, to_csv, to_excel


class TestToCsv:
    """覆盖写 CSV。"""

    def test_writes_header_and_rows(self, tmp_path):
        p = tmp_path / "out.csv"
        to_csv([{"a": "1", "b": "x"}, {"a": "2", "b": "y"}], ["a", "b"], str(p))

        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert rows == [["a", "b"], ["1", "x"], ["2", "y"]]

    def test_numbered_adds_index_column(self, tmp_path):
        p = tmp_path / "out.csv"
        to_csv([{"a": "1"}], ["a"], str(p), numbered=True)

        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert rows[0] == ["序号", "a"]
        assert rows[1] == ["1", "1"]

    def test_dedup_key(self, tmp_path):
        p = tmp_path / "out.csv"
        to_csv([{"link": "L1", "v": 1}, {"link": "L1", "v": 2},
                {"link": "L2", "v": 3}], ["link", "v"], str(p), dedup_key="link")

        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert [r["link"] for r in rows] == ["L1", "L2"]
        assert rows[0]["v"] == "1"  # 保留第一条

    def test_empty_link_rows_not_dropped_when_dedup_key_blank(self, tmp_path):
        """空 link 的行不去重（保持原 xiaohongshu 语义：无链接数据也保留）。"""
        p = tmp_path / "out.csv"
        to_csv([{"link": "", "v": 1}, {"link": "", "v": 2}],
               ["link", "v"], str(p), dedup_key="link")

        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2


class TestToExcel:
    """覆盖写 Excel。"""

    def test_writes_with_sequence_column(self, tmp_path):
        p = tmp_path / "out.xlsx"
        to_excel([{"t": "A", "n": 1}, {"t": "B", "n": 2}],
                 ["t", "n"], ["标题", "数字"], str(p), sheet_title="数据")

        from openpyxl import load_workbook
        ws = load_workbook(p).active
        assert ws.title == "数据"
        assert [c.value for c in ws[1]] == ["序号", "标题", "数字"]
        assert [c.value for c in ws[2]] == [1, "A", 1]


class TestAppendCsv:
    """追加写 + 去重。"""

    def test_first_write_creates_header(self, tmp_path):
        p = tmp_path / "out.csv"
        n = append_csv([{"title": "A", "x": 1}], ["title", "x"], str(p), dedup_key="title")

        assert n == 1
        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["title"] == "A"

    def test_append_dedups_by_key(self, tmp_path):
        p = tmp_path / "out.csv"
        append_csv([{"title": "A"}, {"title": "B"}], ["title"], str(p), dedup_key="title")
        n2 = append_csv([{"title": "A"}, {"title": "C"}], ["title"], str(p), dedup_key="title")

        assert n2 == 1  # A 已存在，只新增 C
        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert [r["title"] for r in rows] == ["A", "B", "C"]

    def test_append_composite_key(self, tmp_path):
        """复合键去重（bilibili 按 title+category）。"""
        p = tmp_path / "out.csv"
        n1 = append_csv([{"title": "T", "category": "tech"},
                         {"title": "T", "category": "game"}],
                        ["title", "category"], str(p), dedup_keys=["title", "category"])
        n2 = append_csv([{"title": "T", "category": "tech"}],
                        ["title", "category"], str(p), dedup_keys=["title", "category"])

        assert n1 == 2
        assert n2 == 0  # 复合键已存在


class TestMergeCsv:
    """合并写 + 覆盖。"""

    def test_merge_overwrites_same_key(self, tmp_path):
        p = tmp_path / "out.csv"
        merge_csv([{"date": "D1", "city": "北京", "t": 1}],
                  ["date", "city", "t"], str(p), key_cols=["date", "city"])
        stats = merge_csv([{"date": "D1", "city": "北京", "t": 99}],
                          ["date", "city", "t"], str(p), key_cols=["date", "city"])

        assert stats == {"total": 1, "new": 0, "updated": 1}
        with open(p, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["t"] == "99"

    def test_merge_preserves_unrelated(self, tmp_path):
        p = tmp_path / "out.csv"
        merge_csv([{"date": "D1", "city": "北京", "t": 1}],
                  ["date", "city", "t"], str(p), key_cols=["date", "city"])
        stats = merge_csv([{"date": "D2", "city": "上海", "t": 2}],
                          ["date", "city", "t"], str(p), key_cols=["date", "city"])

        assert stats["total"] == 2
        assert stats["new"] == 1
